# coding:UTF-8
# import sys
# sys.path.append("C:\Python27\Scripts\PageObject")
from selenium import webdriver
from gloVar import gloVar
from Assert import Assert
from PageObject import PageObject
from Shortcut import Shortcut
from TXT import TXT
import openpyxl
import time
import codecs
import log
import json

class Common:
    def __init__(self):
        gloVar.pathPyFramework = "C:\Python3\Scripts\MJN"
        gloVar.pathPyFrameworkLog = gloVar.pathPyFramework + "\log.txt"
        gloVar.pathPyFrameworkCapture="C:\Python3\Scripts\MJN\Screenshot\\"
        gloVar.pathPyFrameworkCommon = gloVar.pathPyFramework + "\Common"
        gloVar.pathPyFrameworkPageObject = gloVar.pathPyFramework + "\PageObject"
        gloVar.pathPyFrameworkTC = gloVar.pathPyFramework + "\TestCase"
        gloVar.TCResultFile = TXT(gloVar().getLogTCResult())
        gloVar.TCResultFile.write(gloVar.caseName + ":" + "FAIL")
        gloVar.log = log.Logger(gloVar.caseName, gloVar.pathPyFrameworkLog)
        self.driver = webdriver.Chrome()
        gloVar.testResult = True
        gloVar.driver = self.driver
        gloVar.currentWindowHandle = self.driver.current_window_handle
        self.Excel = Excel()
        self.Assert = Assert()
        self.PageObject = PageObject()
        self.Shortcut = Shortcut()
        self.dataJSON = JSON("C:\\Python3\\Scripts\\MJN\\TestData\\TestData.json")

        gloVar.PageObject = self.PageObject
        # self.LoginPage = PageObject.LoginPage()
        # self.Menu = PageObject.Menu()
        # self.DailyWorkPlanPage = PageObject.DailyWorkPlanPage()
        # self.CreatePlanPage = PageObject.CreatePlanPage()
        # self.DailyWorkPlanDetailInformation = PageObject.DailyWorkPlanDetailInformation()

    def get(self, strUrl):
        self.driver.get(strUrl)
        gloVar.log.add("driver.get(" + strUrl + ")")

    def timeOut(self, sec):
        self.driver.implicitly_wait(sec)
        gloVar.log.add("driver.implicitly_wait(" + str(sec) + ")")

    def maximizeWindow(self):
        self.driver.maximize_window()
        gloVar.log.add("driver.maximize_window()")

    def switchWindow(self):
        handles = self.driver.window_handles
        for handle in handles:
            if handle != gloVar.currentWindowHandle: 
                self.driver.switch_to.window(handle)
                self.currentWindow = False
        self.log.add("driver.swtichWindow()")

    def getCurrentURL(self):
        result = self.driver.current_url
        gloVar.log.add("driver.getCurrentURL : " + result)
        return result

    def closeWindow(self):
        gloVar.currentWindowHandle = self.driver.current_window_handle
        self.driver.close()
        gloVar.log.add("driver.closWindow()")

    def sleep(self, sec):
        time.sleep(sec)
        gloVar.log.add("time.sleep(" + str(sec) + ")") 

    def getNowTime():
        return time.strftime("%Y-%m-%d", time.localtime(time.time()))

    def switchAlertAccept(self):
       time.sleep(3)
       a1 = self.driver.switch_to.alert
       time.sleep(1)
       a1.accept()

    def switchAlertDismiss(self):
       time.sleep(3)
       a1 = self.driver.switch_to.alert
       time.sleep(1)
       a1.dismiss()

    def capture(self,pngname):
        if pngname =='':
            pic_path=gloVar.pathPyFrameworkCapture+gloVar.caseName+time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))+'.png'
            self.driver.save_screenshot(pic_path)
            gloVar.log.add("截图保存成功，路径为："+pic_path)
        else:
            pic_path=gloVar.pathPyFrameworkCapture+gloVar.caseName+pngname+'.png'
            self.driver.save_screenshot(pic_path)  
            gloVar.log.add("截图保存成功，路径为："+pic_path)     
    

class Excel:
    def __init__(self):
        self.name = "Excel"

    def workbook(self, path):
        return Workbook(path)

# Excel.Workbook
class Workbook:
    def __init__(self, path):
        try:
            self.wb = openpyxl.load_workbook(filename=path)
            self.path = path
        except FileNotFoundError:#文件不能找到的异常处理
            self.wb = openpyxl.Workbook()
            self.wb.save(path)
            self.path = path
        # else:
        #     print("No Known Exception")
        

    def worksheet(self,shtName):
        return Worksheet(shtName,self.wb,self.path)
    
    def getWorksheetsNames(self):
        return self.wb.sheetnames

# Excel.Workbook.Worksheet
class Worksheet:
    def __init__(self,shtName,wb,path):
        try:
            self.__sht = wb.get_sheet_by_name(shtName)
        except KeyError:
            self.__sht = wb.create_sheet(shtName)
            wb.save(path)
        self.__wb = wb
        self.__path = path
        self.name = self.__sht.title
        self.lastRow = self.__sht.max_row
        self.lastCol = self.__sht.max_column
        self.dic = {}
        for i in range(1, self.lastCol+1):
            self.dic[self.cell(1, i).getValue()] = i
        # print(self.dic)
        
    def cell(self,rowIndex, colIndex):
        if isinstance(colIndex, int):
            return Cell(rowIndex, colIndex, self.__wb, self.__sht, self.__path)
        else:
            colIndexFromDic = self.dic[colIndex]
            return Cell(rowIndex, colIndexFromDic, self.__wb, self.__sht, self.__path)


# Excel.Workbook.Worksheet.Cell
class Cell:
    def __init__(self, rowIndex, colIndex, wb, sht, path):
        self.rowIndex = rowIndex
        self.colIndex = colIndex
        self.__wb = wb
        self.__sht = sht
        self.__path = path
        # super().

    def getValue(self): 
        result = self.__sht.cell(row=self.rowIndex, column=self.colIndex).value
        return result

    def setValue(self, txt):
        self.__sht.cell(row=self.rowIndex,column=self.colIndex, value=txt)
        self.__wb.save(self.__path)

class LogBK:
    def __init__(self, path):
        self.TXT = TXT(path)
        gloVar.log = Log(path)
    def toTXT(self,msg):
        # msg = msg.decode("utf-8", 'ignore')
        print (msg)
        # self.TXT.add(msg)
        logger.loginfo(msg)


class JSON:
    def __init__(self, path):
        self.path = path
        # if not os.path.exists(path):
        #     file = {}

    def read(self):
        fb = open(self.path, 'r')
        data = json.load(fb)
        fb.close()
        return data

    def write(self, data):
        fb = open(self.path, 'w')
        fb.write(json.dumps(data, indent=2))
        fb.close()